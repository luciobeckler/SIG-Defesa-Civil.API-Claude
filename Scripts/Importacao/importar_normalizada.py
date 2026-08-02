"""Gera o import.sql a partir da PLANILHA_NORMALIZADA.xlsx.

Substitui o importar_planilha.py, que lia a planilha crua e refazia a própria
normalização — sem os descartes de tipificação nem os de-para revisados.

Aqui a planilha normalizada é a única fonte: o que estiver nela entra no banco
como está. Nenhuma regra de limpeza mora neste arquivo; ela toda vive em
normalizar_planilha.py, que é onde as decisões foram tomadas e registradas.

Uso:
    python importar_normalizada.py                      # gera out/import.sql
    python importar_normalizada.py --limpar             # apaga ocorrências antes
    python importar_normalizada.py "<planilha.xlsx>"

O SQL é idempotente: reexecutar não duplica (NOT EXISTS / ON CONFLICT).
"""
import collections
import datetime
import re
import sys
from pathlib import Path

import openpyxl

PADRAO = Path(r"C:\Users\lucio\Desktop\TCC\PLANILHA_NORMALIZADA.xlsx")
OUT = Path(__file__).parent / "out"
NA = "N/A"

EMAIL_IMPORTADOR = "importacao@sig.defesacivil.local"
DOMINIO_VISTORIADOR = "vistoriador.importado"


# ── Utilidades SQL ──────────────────────────────────────────────────────────

def vazio(v):
    return v is None or str(v).strip() in ("", NA)


def sql_str(v):
    """Texto ou NULL. N/A da planilha vira NULL de verdade."""
    if vazio(v):
        return "NULL"
    return "'" + str(v).strip().replace("'", "''") + "'"


def sql_str_nn(v, padrao=""):
    """Texto obrigatório: nunca NULL."""
    if vazio(v):
        v = padrao
    return "'" + str(v).strip().replace("'", "''") + "'"


def sql_int(v, padrao="NULL"):
    if vazio(v):
        return padrao
    d = re.sub(r"\D", "", str(v))
    return d if d else padrao


def sql_data(v):
    """dd/mm/aaaa → DATE 'aaaa-mm-dd'."""
    if vazio(v):
        return "NULL"
    s = str(v).strip()
    for f in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return "DATE '" + datetime.datetime.strptime(s, f).date().isoformat() + "'"
        except ValueError:
            pass
    return "NULL"


def sql_ts(data, hora=None):
    """Timestamp no fuso de São Paulo, como o resto do sistema grava."""
    if vazio(data):
        return "NULL"
    s = str(data).strip()
    try:
        d = datetime.datetime.strptime(s, "%d/%m/%Y").date()
    except ValueError:
        return "NULL"
    h = "00:00"
    if not vazio(hora) and re.match(r"^\d{1,2}:\d{2}$", str(hora).strip()):
        h = str(hora).strip()
    return f"(TIMESTAMP '{d.isoformat()} {h}:00' AT TIME ZONE 'America/Sao_Paulo')"


def sql_array(v):
    """'A; B' → ARRAY['A','B']. Vazio → ARRAY[]::text[]."""
    if vazio(v):
        return "ARRAY[]::text[]"
    itens = [i.strip() for i in str(v).split(";") if i.strip() and i.strip() != NA]
    if not itens:
        return "ARRAY[]::text[]"
    return "ARRAY[" + ", ".join("'" + i.replace("'", "''") + "'" for i in itens) + "]"


def sql_bool(v, verdadeiro="SIM"):
    return "TRUE" if str(v).strip().upper() == verdadeiro else "FALSE"


def slug(nome):
    s = re.sub(r"[^a-z0-9]+", ".", str(nome).lower()).strip(".")
    return s or "sem.nome"


# ── Leitura ─────────────────────────────────────────────────────────────────

def ler_planilha(caminho):
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb["OCORRENCIAS"]
    cab = [c.value for c in ws[1]]
    linhas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {cab[i]: row[i] for i in range(len(cab)) if cab[i]}
        if not vazio(d.get("PROTOCOLO")):
            linhas.append(d)

    catalogo = []
    if "CATALOGO A CRIAR" in wb.sheetnames:
        for row in wb["CATALOGO A CRIAR"].iter_rows(min_row=2, values_only=True):
            if row and not vazio(row[0]):
                catalogo.append(str(row[0]).strip())
    return linhas, catalogo


# ── Geração ─────────────────────────────────────────────────────────────────

def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    limpar = "--limpar" in sys.argv[1:]
    xlsx = Path(args[0]) if args else PADRAO

    if not xlsx.exists():
        sys.exit(f"Planilha não encontrada: {xlsx}")

    linhas, catalogo_novo = ler_planilha(xlsx)
    OUT.mkdir(exist_ok=True)

    sql = []
    W = sql.append

    W("-- Importação a partir da PLANILHA NORMALIZADA — importar_normalizada.py")
    W(f"-- Gerado em {datetime.datetime.now().isoformat(timespec='seconds')} | Fonte: {xlsx.name}")
    W("-- Idempotente: reexecutar não duplica (NOT EXISTS / ON CONFLICT).")
    W("BEGIN;")
    W("")

    # ── 0. Limpeza opcional ─────────────────────────────────────────────────
    if limpar:
        W("-- ══════════════════════════════════════════════════════════════════")
        W("-- LIMPEZA (--limpar): apaga ocorrências e dependentes.")
        W("-- USUÁRIOS SÃO PRESERVADOS. Mesma transação: se a carga falhar, volta tudo.")
        W("-- ══════════════════════════════════════════════════════════════════")
        # log_acesso_lgpd referencia ocorrencias sem cascade
        W('DELETE FROM log_acesso_lgpd WHERE "OcorrenciaId" IS NOT NULL;')
        W("DELETE FROM ocorrencias;")
        W("")

    # ── 1. Usuário técnico da importação ────────────────────────────────────
    W("-- Usuário de sistema (autor dos registros importados)")
    W(f"""INSERT INTO usuarios ("Nome","Email","TipoUsuario","Ativo","CriadoEm")
SELECT 'Importação Planilha (sistema)', '{EMAIL_IMPORTADOR}', 'ATENDENTE', FALSE, NOW()
WHERE NOT EXISTS (SELECT 1 FROM usuarios WHERE "Email" = '{EMAIL_IMPORTADOR}');""")
    W("")
    subq_imp = f"(SELECT \"Id\" FROM usuarios WHERE \"Email\" = '{EMAIL_IMPORTADOR}')"

    # ── 2. Contas dos vistoriadores ─────────────────────────────────────────
    vistoriadores = []
    for l in linhas:
        for c in ("VISTORIADOR_1", "VISTORIADOR_2", "VISTORIADOR_3", "VISTORIADOR_4"):
            v = l.get(c)
            if not vazio(v) and str(v).strip() not in vistoriadores:
                vistoriadores.append(str(v).strip())

    W("-- Contas de vistoriadores históricos (desativadas — sem login)")
    for nome in sorted(vistoriadores):
        email = f"v.{slug(nome)}@{DOMINIO_VISTORIADOR}"
        W(f"""INSERT INTO usuarios ("Nome","Email","TipoUsuario","Ativo","CriadoEm")
SELECT {sql_str_nn(nome)}, '{email}', 'VISTORIADOR', FALSE, NOW()
WHERE NOT EXISTS (SELECT 1 FROM usuarios WHERE "Email" = '{email}');""")
    W("")

    def subq_vist(nome):
        return (f"(SELECT \"Id\" FROM usuarios "
                f"WHERE \"Email\" = 'v.{slug(nome)}@{DOMINIO_VISTORIADOR}')")

    # ── 3. Catálogo de opções ───────────────────────────────────────────────
    # As tipificações novas precisam existir antes da carga, senão o sistema
    # não reconhece o valor no select de opções.
    W("-- Opções de catálogo usadas pelos registros importados")
    opcoes = [("TIPIFICACAO", c) for c in catalogo_novo]

    # valores não-enum que a normalização produz nos demais campos
    for campo, coluna in [("GRAU_RISCO", "GRAU_RISCO_ENCONTRADO"),
                          ("INTERDICAO", "INTERDICAO"),
                          ("REGIME_OCUPACAO", "REGIME_OCUPACAO")]:
        vistos = {str(l.get(coluna)).strip() for l in linhas if not vazio(l.get(coluna))}
        for v in sorted(vistos):
            opcoes.append((campo, v))

    for campo, valor in opcoes:
        W(f"""INSERT INTO opcoes_campo_vistoria ("Campo","Valor","Label","CriadoEm")
VALUES ('{campo}', {sql_str_nn(valor)}, {sql_str_nn(valor)}, NOW())
ON CONFLICT ("Campo","Valor") DO NOTHING;""")
    W("")

    # ── 4. Guarda de colisão de protocolo ───────────────────────────────────
    protos = [str(l["PROTOCOLO"]).strip() for l in linhas]
    lista = ", ".join(f"'{p}'" for p in protos)
    W("-- Aborta se algum protocolo já existir e NÃO tiver vindo desta importação.")
    W(f"""DO $$
DECLARE conflitos int;
BEGIN
    SELECT count(*) INTO conflitos
      FROM ocorrencias o
     WHERE o."Protocolo" = ANY (ARRAY[{lista}])
       AND (o."CriadoPorId" IS NULL OR o."CriadoPorId" <> {subq_imp});
    IF conflitos > 0 THEN
        RAISE EXCEPTION 'Importacao abortada: % protocolo(s) ja existem e nao sao desta importacao.', conflitos;
    END IF;
END $$;""")
    W("")

    # ── 5. Ocorrências ──────────────────────────────────────────────────────
    contadores = collections.Counter()

    for l in linhas:
        proto = str(l["PROTOCOLO"]).strip()
        abertura = sql_ts(l.get("DATA_ABERTURA"), l.get("HORA_ABERTURA"))
        subq_oc = f"(SELECT \"Id\" FROM ocorrencias WHERE \"Protocolo\" = '{proto}')"

        W(f"-- ── {proto} " + "─" * 40)

        W(f"""INSERT INTO ocorrencias ("Protocolo","SolicitanteNome","SolicitanteCpf","SolicitanteEmail",
  "SolicitanteTelefone","DescricaoProblema","Status","CriadoPorId","AbertaEm","AtualizadoEm")
SELECT '{proto}', {sql_str_nn(l.get('SOLICITANTE_NOME'), 'Não informado')},
  {sql_str(l.get('SOLICITANTE_CPF'))}, {sql_str(l.get('SOLICITANTE_EMAIL'))},
  {sql_str(l.get('SOLICITANTE_TELEFONE'))}, {sql_str_nn(l.get('DESCRICAO_PROBLEMA'), 'Não informado')},
  '{l.get('STATUS_OCORRENCIA', 'ENCERRADA')}', {subq_imp}, {abertura}, {abertura}
WHERE NOT EXISTS (SELECT 1 FROM ocorrencias WHERE "Protocolo" = '{proto}');""")
        contadores["ocorrencias"] += 1

        # 5a. Localização
        W(f"""INSERT INTO localizacoes ("OcorrenciaId","Endereco","Bairro","Numero","Cidade","Uf")
SELECT {subq_oc}, {sql_str_nn(l.get('ENDERECO'), 'Não informado')},
  {sql_str_nn(l.get('BAIRRO'), 'Não informado')}, {sql_str(l.get('NUMERO'))},
  {sql_str_nn(l.get('CIDADE'), 'Sabará')}, {sql_str_nn(l.get('UF'), 'MG')}
WHERE NOT EXISTS (SELECT 1 FROM localizacoes WHERE "OcorrenciaId" = {subq_oc});""")

        # 5b. Avaliação de risco — só quando há tipificação de verdade.
        #     A tipificação agora é text[]; ver migration TipificacaoInicialMultivalorada.
        tips = l.get("TIPIFICACAO_INICIAL")
        if not vazio(tips):
            W(f"""INSERT INTO avaliacoes_risco ("OcorrenciaId","TipificacaoInicial","GrauRiscoInicial",
  "AbertaPorUsuarioId","Emergencia","RegistradoEm","AtualizadoEm")
SELECT {subq_oc}, {sql_array(tips)},
  {sql_str_nn(l.get('GRAU_RISCO_INICIAL'), 'Não informado')}, {subq_imp},
  {sql_bool(l.get('EMERGENCIA'))}, {abertura}, {abertura}
WHERE NOT EXISTS (SELECT 1 FROM avaliacoes_risco WHERE "OcorrenciaId" = {subq_oc});""")
            contadores["avaliacoes"] += 1

        # 5c. Agendamento + vistoria
        v1 = l.get("VISTORIADOR_1")
        data_vist = l.get("DATA_VISTORIA")
        if not vazio(v1) or not vazio(data_vist):
            status_ag = "CONCLUIDO" if not vazio(data_vist) else "ATIVO"
            subs = [subq_vist(l.get(c)) if not vazio(l.get(c)) else "NULL"
                    for c in ("VISTORIADOR_1", "VISTORIADOR_2",
                              "VISTORIADOR_3", "VISTORIADOR_4")]
            subq_ag = (f"(SELECT \"Id\" FROM agendamentos_vistoria "
                       f"WHERE \"OcorrenciaId\" = {subq_oc} AND \"Numero\" = 1)")
            data_ag = sql_data(l.get("DATA_AGENDAMENTO") if not vazio(l.get("DATA_AGENDAMENTO"))
                               else data_vist)

            W(f"""INSERT INTO agendamentos_vistoria ("OcorrenciaId","Numero","Status","Data",
  "Vistoriador1Id","Vistoriador2Id","Vistoriador3Id","Vistoriador4Id","AgendadoPorId","AgendadoEm")
SELECT {subq_oc}, 1, '{status_ag}', {data_ag},
  {subs[0]}, {subs[1]}, {subs[2]}, {subs[3]}, {subq_imp}, {abertura}
WHERE NOT EXISTS (SELECT 1 FROM agendamentos_vistoria WHERE "OcorrenciaId" = {subq_oc});""")
            contadores["agendamentos"] += 1

            if not vazio(data_vist):
                W(f"""INSERT INTO vistorias ("OcorrenciaId","Numero","AgendamentoId","DataVistoria",
  "HorarioInicio","HorarioTermino","CaracterizacaoDoLocal","Edificacao","Estrutura",
  "NumeroMoradias","NumeroComodos","NumeroPavimentos","NumeroMoradiasNoLote",
  "PossuiUnidadeFamiliar","NumeroAdultos","NumeroCriancas","NumeroIdosos","NumeroDeficientes",
  "TotalMoradores","TipoRisco","GrauRiscoEncontrado","TipificacaoOcorrencia","RegimeOcupacao",
  "Motivacao","AreasAfetadas","Interdicao","Remocao","Orientacoes","Observacoes",
  "EncaminhamentosDeCampo","Vistoriador1Id","Vistoriador2Id","Vistoriador3Id","Vistoriador4Id",
  "RegistradoPorId","RegistradoEm","AtualizadoEm")
SELECT {subq_oc}, 1, {subq_ag}, {sql_data(data_vist)}, INTERVAL '0', INTERVAL '0',
  NULL, 'Não informado', 'Não informado', 0, 0, 0, 0,
  FALSE, 0, 0, 0, 0, {sql_int(l.get('TOTAL_MORADORES'), '0')},
  'Não informado', {sql_str_nn(l.get('GRAU_RISCO_ENCONTRADO'), 'Não informado')},
  {sql_array(l.get('TIPIFICACAO_VISTORIA'))},
  {sql_str_nn(l.get('REGIME_OCUPACAO'), 'Não informado')},
  ARRAY[]::text[], ARRAY[]::text[],
  {sql_str_nn(l.get('INTERDICAO'), 'Não informado')},
  {sql_str_nn(l.get('REMOCAO'), 'Não informado')},
  {sql_array(l.get('ORIENTACOES'))}, {sql_str(l.get('OBSERVACOES_VISTORIA'))},
  ARRAY[]::text[],
  {subs[0]}, {subs[1]}, {subs[2]}, {subs[3]}, {subq_imp}, {abertura}, {abertura}
WHERE {subs[0]} IS NOT NULL
  AND NOT EXISTS (SELECT 1 FROM vistorias WHERE "OcorrenciaId" = {subq_oc});""")
                contadores["vistorias"] += 1

        # 5d. Notificado
        if str(l.get("NOTIFICADO", "")).strip().upper() == "SIM":
            forma = l.get("FORMA_RECEBIMENTO")
            forma = str(forma).strip() if not vazio(forma) else "EMAIL"
            data_notif = (l.get("DATA_RELATORIO") if not vazio(l.get("DATA_RELATORIO"))
                          else l.get("DATA_ABERTURA"))
            W(f"""INSERT INTO notificados ("OcorrenciaId","Nome","DataNotificacao","FormaRecebimento",
  "RegistradoPorId","RegistradoEm")
SELECT {subq_oc}, {sql_str_nn(l.get('SOLICITANTE_NOME'), 'Não informado')},
  {sql_data(data_notif)}, '{forma}', {subq_imp}, {abertura}
WHERE {sql_data(data_notif)} IS NOT NULL
  AND NOT EXISTS (SELECT 1 FROM notificados WHERE "OcorrenciaId" = {subq_oc});""")
            contadores["notificados"] += 1

        W("")

    # ── 6. Sequence de protocolo ────────────────────────────────────────────
    W("-- Evita que novos protocolos do sistema colidam com os importados")
    W("""SELECT setval('seq_protocolo_ano', GREATEST(
    (SELECT last_value FROM seq_protocolo_ano),
    (SELECT COALESCE(MAX(split_part("Protocolo", '-', 2)::int), 0)
       FROM ocorrencias
      WHERE "Protocolo" ~ '^[0-9]{4}-[0-9]+$')
));""")
    W("")
    W("COMMIT;")
    W("")

    destino = OUT / "import.sql"
    destino.write_text("\n".join(sql), encoding="utf-8")

    print(f"Gerado: {destino}  ({destino.stat().st_size/1024/1024:.1f} MB)")
    for k, v in contadores.items():
        print(f"   {k:14s} {v}")
    print(f"   {'vistoriadores':14s} {len(vistoriadores)}")
    print(f"   {'opções catálogo':14s} {len(opcoes)}")
    if limpar:
        print("\n[!] --limpar: o SQL apaga TODAS as ocorrências antes de importar.")


if __name__ == "__main__":
    main()
