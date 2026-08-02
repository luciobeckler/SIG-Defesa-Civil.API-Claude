# -*- coding: utf-8 -*-
"""
Importador da planilha histórica de ocorrências (PLANILHAS DE OCORRENCIAS.xlsx)
para o banco do SIG-Defesa Civil.

Lê as abas OCORRENCIAS_2025 / OCORRENCIAS_2026 (+ enriquecimento da aba
RELATORIOS PRONTOS 2025), normaliza os dados e gera um arquivo SQL idempotente
(out/import.sql) para ser executado no PostgreSQL do servidor.

Uso:
    python importar_planilha.py "C:\\caminho\\PLANILHAS DE OCORRENCIAS.xlsx"

Saídas (pasta out/):
    import.sql            SQL idempotente (BEGIN/COMMIT, ON CONFLICT DO NOTHING)
    previa_ocorrencias.csv  Uma linha por ocorrência com os valores já normalizados
    normalizacoes.csv     Valores brutos -> normalizados (grau, interdição, status)
    rejeitadas.csv        Linhas puladas e o motivo

O de-para de vistoriadores é lido de mapeamento_vistoriadores.csv (edite e
rode de novo). Se o arquivo não existir, uma proposta é gerada automaticamente.
"""

import csv
import re
import sys
import unicodedata
from collections import Counter
from datetime import datetime, date, time
from pathlib import Path

import openpyxl

# ── Configuração ──────────────────────────────────────────────────────────────

TZ = "America/Sao_Paulo"
EMAIL_IMPORTADOR = "importacao@sig.defesacivil.local"
NAO_INFORMADO = "Não informado"
NAO_CONSTATADO = "Não constatado"

BASE = Path(__file__).parent
OUT = BASE / "out"
MAPEAMENTO_CSV = BASE / "mapeamento_vistoriadores.csv"


# ── Helpers de texto ──────────────────────────────────────────────────────────

def sem_acento(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def norm_chave(s) -> str:
    """Chave de comparação: sem acento, maiúscula, espaços colapsados."""
    if s is None:
        return ""
    s = re.sub(r"\s+", " ", str(s)).strip()
    return sem_acento(s).upper()


def limpo(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip()


def sql_str(s) -> str:
    """Literal SQL (escapa aspas simples). Retorna NULL para vazio."""
    s = limpo(s)
    if not s:
        return "NULL"
    return "'" + s.replace("'", "''") + "'"


def sql_str_nn(s, fallback="") -> str:
    """Literal SQL não-nulo (vazio vira fallback)."""
    s = limpo(s) or fallback
    return "'" + s.replace("'", "''") + "'"


def sql_text_array(itens) -> str:
    itens = [limpo(i) for i in itens if limpo(i)]
    if not itens:
        return "ARRAY[]::text[]"
    return "ARRAY[" + ", ".join("'" + i.replace("'", "''") + "'" for i in itens) + "]::text[]"


def titulo(s: str) -> str:
    """Title-case simples preservando siglas curtas."""
    palavras = limpo(s).split(" ")
    out = []
    for p in palavras:
        if len(p) <= 3 and p.isupper():
            out.append(p)
        else:
            out.append(p.capitalize())
    return " ".join(out)


def slug(s: str) -> str:
    s = sem_acento(limpo(s)).lower()
    return re.sub(r"[^a-z0-9]+", ".", s).strip(".")


def so_digitos(s) -> str:
    return re.sub(r"\D", "", str(s or ""))


# ── Datas ─────────────────────────────────────────────────────────────────────

def como_data(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = limpo(v)
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def como_hora(v):
    if v is None or v == "":
        return None
    if isinstance(v, time):
        return v
    if isinstance(v, datetime):
        return v.time()
    s = limpo(v)
    for fmt in ("%H:%M:%S", "%H:%M", "%Hh%M"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            pass
    return None


def sql_timestamptz(d: date, t: time | None) -> str:
    t = t or time(8, 0)
    return f"(TIMESTAMP '{d.isoformat()} {t.strftime('%H:%M:%S')}' AT TIME ZONE '{TZ}')"


# ── Normalizações de domínio ──────────────────────────────────────────────────

def norm_grau(v) -> str:
    k = norm_chave(v)
    if k in ("ALTO",):
        return "ALTO"
    if k in ("BAIXO",):
        return "BAIXO"
    if k in ("MEDIO",):
        return "MEDIO"
    if k in ("MUITO ALTO",):
        return "MUITO_ALTO"
    if k in ("NAO CONSTATADO", "NAO ATESTADO", "NAO CONTATADO"):
        return NAO_CONSTATADO
    # NAO INFORMADO / NAO ESPECIFICADO / PENDENTE / '-' / '****' / vazio / demais
    return NAO_INFORMADO


def norm_interdicao(v) -> str:
    k = norm_chave(v)
    if k in ("", "NAO INFORMADO"):
        return NAO_INFORMADO
    if k.startswith("NAO") or k == "N":
        return "NAO_NECESSARIA"
    if "PARCIAL" in k:
        return "PARCIAL"
    if k == "SIM":
        return "TOTAL"
    return titulo(v)  # DESOCUPAÇÃO, SINALIZAÇÃO etc. — mantém como texto


def norm_simnao(v) -> str | None:
    k = norm_chave(v)
    if k.startswith("S"):
        return "SIM"
    if k.startswith("N"):
        return "NÃO"
    return None


def derivar_status(status_vistoria, status_relatorio) -> tuple[str, str]:
    """Retorna (status_ocorrencia, modo) — modo: REALIZADA | PENDENTE | DISPENSAVEL."""
    sv = norm_chave(status_vistoria)
    sr = norm_chave(status_relatorio)
    if sv == "REALIZADA":
        if sr in ("CONCLUIDO", "DISPENSAVEL"):
            return "ENCERRADA", "REALIZADA"
        return "VISTORIA_REALIZADA", "REALIZADA"
    if sv == "DISPENSAVEL":
        return "ENCERRADA", "DISPENSAVEL"
    return "VISTORIA_SOLICITADA", "PENDENTE"


def parse_vistoriadores(v) -> list[str]:
    if not v:
        return []
    s = limpo(v)
    s = re.sub(r"\bE\b", ",", s, flags=re.IGNORECASE)
    partes = re.split(r"[,/;]+", s)
    return [p.strip() for p in partes if p.strip() and norm_chave(p) not in ("", "-")]


# ── De-para de vistoriadores ──────────────────────────────────────────────────

def carregar_ou_propor_mapeamento(nomes_brutos: Counter) -> dict[str, str]:
    """Lê mapeamento_vistoriadores.csv (nome_planilha;nome_canonico).
    Se não existir, gera proposta (merge apenas por acento/caixa) e usa-a."""
    if MAPEAMENTO_CSV.exists():
        mapa = {}
        with open(MAPEAMENTO_CSV, encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f, delimiter=";"):
                mapa[norm_chave(row["nome_planilha"])] = limpo(row["nome_canonico"])
        return mapa

    # Proposta: agrupa por chave sem acento; canônico = forma mais frequente em title-case
    grupos: dict[str, Counter] = {}
    for nome, qtd in nomes_brutos.items():
        grupos.setdefault(norm_chave(nome), Counter())[limpo(nome)] += qtd
    mapa = {}
    linhas = []
    for chave, formas in sorted(grupos.items()):
        canonico = titulo(formas.most_common(1)[0][0])
        mapa[chave] = canonico
        total = sum(formas.values())
        linhas.append((formas.most_common(1)[0][0], canonico, total))
    with open(MAPEAMENTO_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["nome_planilha", "nome_canonico", "ocorrencias_na_planilha"])
        for bruto, canonico, total in sorted(linhas, key=lambda x: -x[2]):
            w.writerow([bruto, canonico, total])
    print(f"[!] Proposta de de-para gerada em {MAPEAMENTO_CSV.name} — revise os nomes")
    print("    (ex.: unifique 'Rogerio'/'Paulo Rogerio' se forem a mesma pessoa) e rode de novo.")
    return mapa


# ── Leitura da planilha ───────────────────────────────────────────────────────

def ler_aba(ws) -> list[dict]:
    hdr = [limpo(c.value) for c in ws[1]]
    linhas = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        d = {hdr[i]: r[i] for i in range(len(hdr)) if hdr[i]}
        linhas.append(d)
    return linhas


def num_vistoria(v) -> str:
    """2025001.0 -> '2025001'"""
    s = limpo(v)
    if s.endswith(".0"):
        s = s[:-2]
    return so_digitos(s)


def formatar_protocolo(numero: str) -> str | None:
    """
    Converte o número da planilha para o formato de protocolo do sistema:
        2025001 -> 2025-0001   |   2026653 -> 2026-0653
    Os 4 primeiros dígitos são o ano; o restante é a sequência (4 dígitos).
    Retorna None se o número não tiver o formato esperado.
    """
    if len(numero) < 5 or not numero.isdigit():
        return None
    ano, seq = numero[:4], numero[4:]
    return f"{ano}-{int(seq):04d}"


# ── Geração do SQL ────────────────────────────────────────────────────────────

def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    flags = {a for a in sys.argv[1:] if a.startswith("--")}
    limpar = "--limpar" in flags

    xlsx = args[0] if args else r"C:\Users\lucio\Downloads\PLANILHAS DE OCORRENCIAS.xlsx"
    OUT.mkdir(exist_ok=True)

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    l2025 = ler_aba(wb["OCORRENCIAS_2025"])
    l2026 = ler_aba(wb["OCORRENCIAS_2026"])

    # Enriquecimento 2025: aba de relatórios indexada pelo nº da ocorrência
    rel2025 = {}
    if "RELATORIOS PRONTOS 2025" in wb.sheetnames:
        for r in ler_aba(wb["RELATORIOS PRONTOS 2025"]):
            chave = num_vistoria(r.get("Nº OCORRENCIA") or r.get("N° OCORRENCIA") or "")
            if chave:
                rel2025[chave] = r

    todas = [("2025", l) for l in l2025] + [("2026", l) for l in l2026]

    # Coleta nomes de vistoriadores para o de-para
    nomes = Counter()
    for _, l in todas:
        for n in parse_vistoriadores(l.get("VISTORIADORES")):
            nomes[n] += 1
    mapa_vist = carregar_ou_propor_mapeamento(nomes)

    def canonico(nome: str) -> str:
        """Nome canônico do de-para. Valor em branco no CSV = ignorar o nome
        (ex.: 'DEMAIS SECRETARIAS', que não é uma pessoa)."""
        chave = norm_chave(nome)
        if chave in mapa_vist:
            return mapa_vist[chave]  # pode ser "" → ignorado
        return titulo(nome)

    sql: list[str] = []
    previa, normalizacoes, rejeitadas = [], set(), []
    vistoriadores_canonicos: set[str] = set()

    W = sql.append
    W("-- Importação da planilha histórica de ocorrências — gerado por importar_planilha.py")
    W(f"-- Gerado em {datetime.now().isoformat(timespec='seconds')} | Fonte: {Path(xlsx).name}")
    W("-- Idempotente: reexecutar não duplica registros (ON CONFLICT / NOT EXISTS).")
    W("BEGIN;")
    W("")

    # 0. Limpeza da base de testes (opcional, via --limpar).
    #    Remove SOMENTE as ocorrências e tudo que depende delas. Os usuários
    #    (admin, servidores) são preservados. Roda na MESMA transação: se a
    #    importação falhar depois, a limpeza também é desfeita.
    if limpar:
        W("-- ══════════════════════════════════════════════════════════════════")
        W("-- LIMPEZA DA BASE DE TESTES (--limpar)")
        W("-- Apaga as ocorrências e seus dependentes. USUÁRIOS SÃO PRESERVADOS.")
        W("-- ══════════════════════════════════════════════════════════════════")
        # log_acesso_lgpd referencia ocorrencias com NO ACTION → apagar antes.
        W('DELETE FROM log_acesso_lgpd WHERE "OcorrenciaId" IS NOT NULL;')
        # As demais dependentes (localizacoes, arquivos, avaliacoes_risco,
        # agendamentos_vistoria -> tentativas, vistorias, notificados,
        # encaminhamentos_finais, Observacoes) caem por CASCADE.
        W("DELETE FROM ocorrencias;")
        # Catálogo de opções personalizadas criado durante os testes.
        W("DELETE FROM opcoes_campo_vistoria;")
        W("")

    # 1. Usuário de sistema para a importação
    W("-- Usuário de sistema (autor dos registros importados)")
    W(f"""INSERT INTO usuarios ("Nome","Email","TipoUsuario","Ativo","CriadoEm")
SELECT 'Importação Planilha (sistema)', '{EMAIL_IMPORTADOR}', 'ATENDENTE', FALSE, NOW()
WHERE NOT EXISTS (SELECT 1 FROM usuarios WHERE "Email" = '{EMAIL_IMPORTADOR}');""")
    W("")

    # 2. Catálogo: opções não-enum usadas pela importação
    W("-- Opções personalizadas usadas pelos registros históricos")
    for campo, valor in [
        ("GRAU_RISCO", NAO_INFORMADO), ("GRAU_RISCO", NAO_CONSTATADO),
        ("TIPO_RISCO", NAO_INFORMADO), ("EDIFICACAO", NAO_INFORMADO),
        ("ESTRUTURA", NAO_INFORMADO), ("REGIME_OCUPACAO", NAO_INFORMADO),
        ("AREA_AFETADA", NAO_INFORMADO), ("REMOCAO", NAO_INFORMADO),
        ("INTERDICAO", NAO_INFORMADO), ("TIPIFICACAO", NAO_INFORMADO),
    ]:
        W(f"""INSERT INTO opcoes_campo_vistoria ("Campo","Valor","Label","CriadoEm")
VALUES ('{campo}', {sql_str_nn(valor)}, {sql_str_nn(valor)}, NOW())
ON CONFLICT ("Campo","Valor") DO NOTHING;""")
    W("")

    subq_importador = f"(SELECT \"Id\" FROM usuarios WHERE \"Email\" = '{EMAIL_IMPORTADOR}')"

    def subq_vistoriador(nome_canonico: str) -> str:
        return f"(SELECT \"Id\" FROM usuarios WHERE \"Email\" = 'v.{slug(nome_canonico)}@vistoriador.importado')"

    # 3. Linha a linha
    protocolos_gerados: dict[str, str] = {}  # protocolo -> número original (detecta colisão)

    for ano, l in todas:
        num_planilha = num_vistoria(l.get("N_DA_VISTORIA"))
        proto = formatar_protocolo(num_planilha)
        if not proto:
            rejeitadas.append({"protocolo": limpo(l.get('N_DA_VISTORIA')), "motivo": "número de vistoria inválido"})
            continue

        if proto in protocolos_gerados:
            rejeitadas.append({
                "protocolo": proto,
                "motivo": f"protocolo duplicado (números {protocolos_gerados[proto]} e {num_planilha} geram o mesmo)",
            })
            continue
        protocolos_gerados[proto] = num_planilha

        nome_sol = limpo(l.get("NOME DO SOLICITANTE")) or "Não informado"
        data_sol = como_data(l.get("DATA_SOLICITACAO"))
        if not data_sol:
            rejeitadas.append({"protocolo": proto, "motivo": "sem DATA_SOLICITACAO"})
            continue
        hora_sol = como_hora(l.get("HORARIO"))

        cpf = so_digitos(l.get("CPF/IDENTIDADE"))
        cpf = cpf if len(cpf) == 11 and cpf != "0" * 11 else ""
        telefone = limpo(l.get("TELEFONE"))
        email_l = limpo(l.get("EMAIL")) if ano == "2026" else ""
        endereco = limpo(l.get("ENDEREÇO")) or "Não informado"
        numero = num_vistoria(l.get("Nº")) or limpo(l.get("Nº"))
        bairro = titulo(l.get("BAIRRO")) or "Não informado"

        tipif = limpo(l.get("TIPIFICACAO_OCORRENCIA")) or NAO_INFORMADO
        descricao = limpo(l.get("DESCRIÇÃO PRELIMINAR")) or tipif

        grau = norm_grau(l.get("GRAU_RISCO"))
        interdicao = norm_interdicao(l.get("INTERDIÇÃO"))
        notif = norm_simnao(l.get("NOTIFICAÇÃO"))
        status, modo = derivar_status(l.get("STATUS_VISTORIA"), l.get("STATUS_RELATORIO"))

        normalizacoes.add(("GRAU_RISCO", limpo(l.get("GRAU_RISCO")), grau))
        normalizacoes.add(("INTERDIÇÃO", limpo(l.get("INTERDIÇÃO")), interdicao))
        normalizacoes.add(("STATUS", f"{limpo(l.get('STATUS_VISTORIA'))} + {limpo(l.get('STATUS_RELATORIO'))}", status))

        data_vist = como_data(l.get("DATA DA VISTORIA")) or data_sol
        equipe = [canonico(n) for n in parse_vistoriadores(l.get("VISTORIADORES"))]
        # remove ignorados (canônico em branco) e dedupe preservando ordem
        equipe = list(dict.fromkeys(n for n in equipe if n))
        vistoriadores_canonicos.update(equipe)
        # Equipe de até 4 pessoas (colunas Vistoriador1..4); excedentes vão p/ observação
        v1, v2, v3, v4 = (equipe + [None] * 4)[:4]
        extras = equipe[4:]

        # Observações consolidadas
        obs_partes = []
        for rotulo, col in [("Obs", "OBSERVAÇÃO"), ("Despachos enviados", "DESPACHOS ENVIADOS"),
                            ("Resposta despachos", "RESPOSTA DESPACHOS")]:
            valor = limpo(l.get(col))
            if valor:
                obs_partes.append(f"{rotulo}: {valor}")
        if notif == "SIM":
            obs_partes.append("Notificação emitida: SIM")
        if extras:
            obs_partes.append("Equipe adicional: " + ", ".join(extras))
        dr = como_data(l.get("DATA_RELATORIO"))
        if dr:
            obs_partes.append(f"Relatório: {norm_chave(l.get('STATUS_RELATORIO'))} em {dr.strftime('%d/%m/%Y')}")
        else:
            sr = limpo(l.get("STATUS_RELATORIO"))
            if sr:
                obs_partes.append(f"Relatório: {norm_chave(sr)}")
        if modo == "DISPENSAVEL":
            obs_partes.append("Vistoria dispensável (registro histórico sem vistoria)")

        # Enriquecimento 2025 (aba de relatórios)
        total_moradores, regime, orientacoes, encaminhamentos = 0, NAO_INFORMADO, [], []
        rel = rel2025.get(proto)
        if rel:
            tm = so_digitos(rel.get("Nº DE MORADORES") or rel.get("N° DE MORADORES"))
            total_moradores = int(tm) if tm else 0
            ocup = limpo(rel.get("OCUPAÇÃO DO IMOVEL"))
            if ocup:
                regime = titulo(ocup)
            ori = limpo(rel.get("ORIENTAÇÕES"))
            if ori:
                orientacoes = [ori]
            enc = limpo(rel.get("ENCAMINHAMENTOS"))
            if enc:
                encaminhamentos = [e.strip() for e in enc.split(",") if e.strip()]
            conclusao = limpo(rel.get("CONCLUSÃO"))
            if conclusao:
                obs_partes.append(f"Conclusão do relatório: {conclusao}")
            entrega = limpo(rel.get("FORMA DE ENTREGA DO RELATORIO"))
            if entrega:
                obs_partes.append(f"Entrega do relatório: {entrega}")

        observacoes = " | ".join(obs_partes)
        abertura_ts = sql_timestamptz(data_sol, hora_sol)

        W(f"-- ── Ocorrência {proto} ({status}) " + "─" * 30)

        # 3a. Solicitante — gravado na própria ocorrência.
        #     Cidadãos não são usuários do sistema (não têm conta nem login).
        sql_cpf = f"'{cpf}'" if cpf else "NULL"
        email_sol = email_l or (f"cid.{cpf}@importado.local" if cpf
                                else f"sol.{num_planilha}@importado.local")

        # 3b. Ocorrência
        W(f"""INSERT INTO ocorrencias ("Protocolo","SolicitanteNome","SolicitanteCpf","SolicitanteEmail","SolicitanteTelefone","DescricaoProblema","Status","CriadoPorId","AbertaEm","AtualizadoEm")
SELECT '{proto}', {sql_str_nn(nome_sol)}, {sql_cpf}, {sql_str_nn(email_sol)}, {sql_str(telefone)}, {sql_str_nn(descricao)}, '{status}', {subq_importador}, {abertura_ts}, {abertura_ts}
WHERE NOT EXISTS (SELECT 1 FROM ocorrencias WHERE "Protocolo" = '{proto}');""")
        subq_oc = f"(SELECT \"Id\" FROM ocorrencias WHERE \"Protocolo\" = '{proto}')"

        # 3c. Localização
        W(f"""INSERT INTO localizacoes ("OcorrenciaId","Endereco","Bairro","Numero","Cidade","Uf")
SELECT {subq_oc}, {sql_str_nn(endereco)}, {sql_str_nn(bairro)}, {sql_str(numero)}, 'Sabará', 'MG'
WHERE NOT EXISTS (SELECT 1 FROM localizacoes WHERE "OcorrenciaId" = {subq_oc});""")

        # 3d. Agendamento (+ tentativa) — quando há fluxo de vistoria
        if modo in ("REALIZADA", "PENDENTE"):
            status_ag = "CONCLUIDO" if modo == "REALIZADA" else "ATIVO"
            v1_sub = subq_vistoriador(v1) if v1 else "NULL"
            v2_sub = subq_vistoriador(v2) if v2 else "NULL"
            v3_sub = subq_vistoriador(v3) if v3 else "NULL"
            v4_sub = subq_vistoriador(v4) if v4 else "NULL"
            W(f"""INSERT INTO agendamentos_vistoria ("OcorrenciaId","Numero","Status","Data","Vistoriador1Id","Vistoriador2Id","Vistoriador3Id","Vistoriador4Id","AgendadoPorId","AgendadoEm")
SELECT {subq_oc}, 1, '{status_ag}', DATE '{data_vist.isoformat()}', {v1_sub}, {v2_sub}, {v3_sub}, {v4_sub}, {subq_importador}, {abertura_ts}
WHERE NOT EXISTS (SELECT 1 FROM agendamentos_vistoria WHERE "OcorrenciaId" = {subq_oc});""")
            subq_ag = f"(SELECT \"Id\" FROM agendamentos_vistoria WHERE \"OcorrenciaId\" = {subq_oc} AND \"Numero\" = 1)"
            W(f"""INSERT INTO tentativas_vistoria ("AgendamentoId","NumeroTentativa","DataHoraTentativa","RegistradoEm")
SELECT {subq_ag}, 1, {sql_timestamptz(data_vist, time(8, 0))}, {abertura_ts}
WHERE NOT EXISTS (SELECT 1 FROM tentativas_vistoria WHERE "AgendamentoId" = {subq_ag});""")

        # 3e. Vistoria — apenas para REALIZADA
        if modo == "REALIZADA":
            v1_final = subq_vistoriador(v1) if v1 else subq_importador
            v2_final = subq_vistoriador(v2) if v2 else "NULL"
            v3_final = subq_vistoriador(v3) if v3 else "NULL"
            v4_final = subq_vistoriador(v4) if v4 else "NULL"
            registrado_ts = sql_timestamptz(data_vist, time(17, 0))
            W(f"""INSERT INTO vistorias ("OcorrenciaId","Numero","AgendamentoId","DataVistoria","HorarioInicio","HorarioTermino",
  "CaracterizacaoDoLocal","Edificacao","Estrutura","NumeroMoradias","NumeroComodos","NumeroPavimentos","NumeroMoradiasNoLote",
  "PossuiUnidadeFamiliar","NumeroAdultos","NumeroCriancas","NumeroIdosos","NumeroDeficientes","TotalMoradores",
  "TipoRisco","GrauRiscoEncontrado","TipificacaoOcorrencia","RegimeOcupacao","Motivacao","AreasAfetadas",
  "Interdicao","Remocao","Orientacoes","Observacoes","EncaminhamentosDeCampo",
  "Vistoriador1Id","Vistoriador2Id","Vistoriador3Id","Vistoriador4Id","RegistradoPorId","RegistradoEm","AtualizadoEm")
SELECT {subq_oc}, 1, {subq_ag}, DATE '{data_vist.isoformat()}', INTERVAL '0', INTERVAL '0',
  NULL, {sql_str_nn(NAO_INFORMADO)}, {sql_str_nn(NAO_INFORMADO)}, 0, 0, 0, 0,
  FALSE, 0, 0, 0, 0, {total_moradores},
  {sql_str_nn(NAO_INFORMADO)}, {sql_str_nn(grau)}, {sql_text_array([tipif])}, {sql_str_nn(regime)}, ARRAY[]::text[], {sql_text_array([NAO_INFORMADO])},
  {sql_str_nn(interdicao)}, {sql_str_nn(NAO_INFORMADO)}, {sql_text_array(orientacoes)}, {sql_str(observacoes)}, {sql_text_array(encaminhamentos)},
  {v1_final}, {v2_final}, {v3_final}, {v4_final}, {subq_importador}, {registrado_ts}, {registrado_ts}
WHERE NOT EXISTS (SELECT 1 FROM vistorias WHERE "OcorrenciaId" = {subq_oc});""")
        W("")

        previa.append({
            "protocolo": proto, "numero_planilha": num_planilha, "ano": ano, "status": status, "solicitante": nome_sol,
            "cpf": cpf or "-", "endereco": f"{endereco}, {numero}", "bairro": bairro,
            "aberta_em": f"{data_sol} {hora_sol or ''}".strip(), "data_vistoria": data_vist if modo != "DISPENSAVEL" else "-",
            "grau_risco": grau if modo == "REALIZADA" else "-", "interdicao": interdicao if modo == "REALIZADA" else "-",
            "vistoriador_1": v1 or "-", "vistoriador_2": v2 or "-",
            "vistoriador_3": v3 or "-", "vistoriador_4": v4 or "-",
            "equipe_extra": ", ".join(extras) or "-", "observacoes": observacoes[:200],
        })

    # 4. Contas dos vistoriadores (desativadas — apenas para exibição de nome)
    bloco_vist = ["-- Contas de vistoriadores históricos (desativadas — sem login)"]
    for nome in sorted(vistoriadores_canonicos):
        email = f"v.{slug(nome)}@vistoriador.importado"
        bloco_vist.append(f"""INSERT INTO usuarios ("Nome","Email","TipoUsuario","Ativo","CriadoEm")
SELECT {sql_str_nn(nome)}, '{email}', 'VISTORIADOR', FALSE, NOW()
WHERE NOT EXISTS (SELECT 1 FROM usuarios WHERE "Email" = '{email}');""")
    bloco_vist.append("")

    # 4b. Guarda de colisão de protocolo.
    # Com o formato AAAA-NNNN, um protocolo da planilha pode coincidir com um
    # gerado pelo próprio sistema. Nesse caso a ocorrência seria pulada, mas o
    # agendamento/vistoria acabaria pendurado na ocorrência errada. Abortamos.
    # Protocolos criados por uma execução anterior DESTA importação são aceitos
    # (mantém a idempotência).
    lista_protos = ",\n        ".join(
        "'" + p + "'" for p in sorted(protocolos_gerados)
    )
    bloco_guard = [
        "-- ── Guarda: colisão de protocolo ────────────────────────────────────",
        "DO $$",
        "DECLARE conflitos int;",
        "BEGIN",
        "    SELECT count(*) INTO conflitos",
        "      FROM ocorrencias o",
        "     WHERE o.\"Protocolo\" = ANY (ARRAY[",
        f"        {lista_protos}",
        "     ])",
        f"       AND o.\"CriadoPorId\" <> (SELECT \"Id\" FROM usuarios WHERE \"Email\" = '{EMAIL_IMPORTADOR}');",
        "    IF conflitos > 0 THEN",
        "        RAISE EXCEPTION 'Importacao abortada: % protocolo(s) da planilha ja existem no banco e nao foram criados por esta importacao. Resolva os conflitos antes de importar.', conflitos;",
        "    END IF;",
        "END $$;",
        "",
    ]

    # Insere guarda + vistoriadores logo após o catálogo (antes das ocorrências)
    pos = sql.index("")  # primeira linha em branco após o cabeçalho
    # encontra o fim do bloco de catálogo (última linha antes da primeira ocorrência)
    for i, linha in enumerate(sql):
        if linha.startswith("-- ── Ocorrência"):
            pos = i
            break
    sql[pos:pos] = bloco_guard + bloco_vist

    # 5. Avança a sequence de protocolo para além dos números importados,
    #    senão o sistema emitiria protocolos já usados (ex.: 2026-0007).
    W("-- ── Sequence de protocolo ────────────────────────────────────────────")
    W("-- Garante que novos protocolos gerados pelo sistema não colidam com os importados.")
    W("""SELECT setval('seq_protocolo_ano', GREATEST(
    (SELECT last_value FROM seq_protocolo_ano),
    (SELECT COALESCE(MAX(split_part("Protocolo", '-', 2)::int), 0)
       FROM ocorrencias
      WHERE "Protocolo" ~ '^[0-9]{4}-[0-9]+$')
));""")
    W("")

    W("COMMIT;")

    # ── Arquivos de saída ────────────────────────────────────────────────────
    (OUT / "import.sql").write_text("\n".join(sql), encoding="utf-8")

    with open(OUT / "previa_ocorrencias.csv", "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(previa[0].keys()), delimiter=";")
        w.writeheader()
        w.writerows(previa)

    with open(OUT / "normalizacoes.csv", "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["campo", "valor_planilha", "valor_importado"])
        for linha in sorted(normalizacoes):
            w.writerow(linha)

    with open(OUT / "rejeitadas.csv", "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["protocolo", "motivo"], delimiter=";")
        w.writeheader()
        w.writerows(rejeitadas)

    if limpar:
        print("[!] MODO --limpar: o import.sql apaga TODAS as ocorrências existentes")
        print("    (e o catálogo de opções) antes de importar. Usuários são preservados.")
    print(f"[ok] {len(previa)} ocorrências geradas | {len(rejeitadas)} rejeitadas")
    print(f"[ok] Vistoriadores (contas desativadas): {len(vistoriadores_canonicos)}")
    print(f"[ok] Saídas em {OUT}")


if __name__ == "__main__":
    main()
