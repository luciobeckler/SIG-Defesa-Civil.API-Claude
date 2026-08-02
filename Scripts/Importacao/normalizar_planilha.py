"""Normaliza a planilha histórica de ocorrências para o formato do sistema.

Saída: PLANILHA_NORMALIZADA.xlsx, com a aba de dados no formato que será
inserido no sistema e abas de conferência (de-para e pendências), para o usuário
revisar cada decisão antes de importar.

Princípio: quando a decisão exige conhecimento que o dado não carrega — se
"Leandro" é o Leandro Santos ou o Leandro de Jesus —, o script NÃO adivinha.
Mantém separado e marca para revisão. Unir duas pessoas por engano é pior do que
deixar duas linhas para conferir.
"""
import collections, datetime, re, unicodedata
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ENTRADA = r"C:\Users\lucio\Desktop\TCC\PLANILHAS DE OCORRENCIAS (1).xlsx"
SAIDA = r"C:\Users\lucio\Desktop\TCC\PLANILHA_NORMALIZADA.xlsx"
NA = "N/A"

# ════════════════════════════════════════════════════════════════════════════
#  Utilidades
# ════════════════════════════════════════════════════════════════════════════

def texto(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return " ".join(str(v).split())


def chave(s):
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    return " ".join(s.upper().split())


def titulo(s):
    """Capitaliza preservando siglas curtas e preposições em minúscula."""
    s = texto(s)
    if not s:
        return ""
    miu = {"DE", "DA", "DO", "DAS", "DOS", "E"}
    partes = []
    for i, p in enumerate(s.split()):
        u = chave(p)
        if u in miu and i > 0:
            partes.append(p.lower())
        elif len(p) <= 3 and p.isupper() and u not in miu:
            partes.append(p)             # sigla: SÃO, BH, SGT
        else:
            partes.append(p.capitalize())
    return " ".join(partes)


def so_digitos(v):
    return re.sub(r"\D", "", texto(v))


def como_data(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    s = texto(v)
    for f in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(s, f).date()
        except ValueError:
            pass
    return None


def como_hora(v):
    if isinstance(v, datetime.time):
        return v.strftime("%H:%M")
    if isinstance(v, datetime.datetime):
        return v.strftime("%H:%M")
    s = texto(v)
    m = re.match(r"^(\d{1,2})[:hH](\d{2})", s)
    return f"{int(m.group(1)):02d}:{m.group(2)}" if m else ""


# ════════════════════════════════════════════════════════════════════════════
#  Catálogo do sistema (frontend/src/app/shared/utils/enum-options.ts)
# ════════════════════════════════════════════════════════════════════════════

TIPIFICACOES_SISTEMA = [
    "ABATIMENTO_DE_FOSSA", "ALAGAMENTO", "ARVORE_COM_RISCO_DE_QUEDA",
    "CICATRIZ_DE_ESCORREGAMENTO", "DEGRAU_DE_ABATIMENTO", "EROSAO",
    "ESCORREGAMENTO", "INCENDIO", "INUNDACAO_DE_CORREGO_RIO",
    "QUEDA_DE_ARVORES", "REDE_PUBLICA_DE_DRENAGEM_PLUVIAL_ROMPIDA",
    "ROLAMENTO_TOMBAMENTO_DE_BLOCOS", "SOLAPAMENTO", "TRINCAS",
]
GRAUS_SISTEMA = ["BAIXO", "MEDIO", "ALTO", "MUITO_ALTO"]

# Termo da planilha → código do sistema.
# "novo" = precisa ser criado no catálogo (OpcaoCampoVistoria) antes de importar.
MAPA_TIPIFICACAO = {
    # já existem no sistema
    "ESCORREGAMENTO": ("ESCORREGAMENTO", False),
    "ESCOREGAMENTO": ("ESCORREGAMENTO", False),
    "DESLIZAMENTO": ("ESCORREGAMENTO", False),
    "DESLIZAMENTO DE TERRA": ("ESCORREGAMENTO", False),
    "RISCO DE DESLIZAMENTO": ("ESCORREGAMENTO", False),
    "TRINCAS": ("TRINCAS", False),
    "EROSAO": ("EROSAO", False),
    "EROSAO EM VIA PUBLICA": ("EROSAO", False),
    "ALAGAMENTO": ("ALAGAMENTO", False),
    "INCENDIO": ("INCENDIO", False),
    "INCENDIO EM RESIDENCIA": ("INCENDIO", False),
    "INCENDIO EM EDIFICACAO": ("INCENDIO", False),
    "QUEDA DE ARVORE": ("QUEDA_DE_ARVORES", False),
    "QUEDA DE ARVORES": ("QUEDA_DE_ARVORES", False),
    "ARVORE COM RISCO DE QUEDA": ("ARVORE_COM_RISCO_DE_QUEDA", False),
    "SOLAPAMENTO": ("SOLAPAMENTO", False),
    "INUNDACAO": ("INUNDACAO_DE_CORREGO_RIO", False),
    "ABATIMENTO DE FOSSA": ("ABATIMENTO_DE_FOSSA", False),
    "INEXIST./INSULF. DE DRENAGEM PLUVIAL": ("REDE_PUBLICA_DE_DRENAGEM_PLUVIAL_ROMPIDA", False),
    "INSUFICIENCIA DO SISTEMA DE CAPTACAO E CONDUCAO DE AGUAS PLUVIAIS":
        ("REDE_PUBLICA_DE_DRENAGEM_PLUVIAL_ROMPIDA", False),

    # não existem — precisam entrar no catálogo
    "AVALIACAO DE RISCO": ("AVALIACAO_DE_RISCO", True),
    "AVALIACOA DE RISCO": ("AVALIACAO_DE_RISCO", True),
    "FISSURAS": ("FISSURAS", True),
    "RACHADURAS": ("RACHADURAS", True),
    "INFILTRACAO": ("INFILTRACAO", True),
    "MURO DE ARRIMO/DIVISA COM ANOMALIAS": ("MURO_DE_ARRIMO_COM_ANOMALIAS", True),
    "COLAPSO DE MURO DE ARRIMO/DIVISA": ("COLAPSO_DE_MURO_DE_ARRIMO", True),
    "DESABAMENTO TOTAL MURO": ("COLAPSO_DE_MURO_DE_ARRIMO", True),
    "DESLOCAMENTO MURO": ("MURO_DE_ARRIMO_COM_ANOMALIAS", True),
    "DESTELHAMENTO": ("DESTELHAMENTO", True),
    "DESABAMENTO PARCIAL": ("DESABAMENTO_PARCIAL", True),
    "DESABAMENTO": ("DESABAMENTO_PARCIAL", True),
    "DESPRENDIMENTO DE REBOCO": ("DESPRENDIMENTO_DE_REBOCO", True),
    "LANCAMENTO DE AGUA PLUVIAL/ESGOTO": ("LANCAMENTO_IRREGULAR_DE_AGUA", True),
    "RESPOSTA DE EMERGENCIA": ("RESPOSTA_DE_EMERGENCIA", True),
    "EMERGENCIA": ("RESPOSTA_DE_EMERGENCIA", True),
    "VISTORIA CAUTELAR": ("VISTORIA_CAUTELAR", True),
    "CADASTRO HABITACIONAL": ("CADASTRO_HABITACIONAL", True),
    "INVASAO": ("INVASAO", True),
    "DESLOCAMENTO": ("DESLOCAMENTO_DE_ESTRUTURA", True),

    # já existem no sistema, grafados por extenso na planilha
    "CICATRIZ DE ESCORREGAMENTO": ("CICATRIZ_DE_ESCORREGAMENTO", False),
    "DEGRAU DE ABATIMENTO": ("DEGRAU_DE_ABATIMENTO", False),
    "AVALIACAO EM INDIVIDUO ARBOREO": ("ARVORE_COM_RISCO_DE_QUEDA", False),
    "ARVORE": ("ARVORE_COM_RISCO_DE_QUEDA", False),
    "INSUFICIENCIA DO SISTEMA DE DRENAGEM": ("REDE_PUBLICA_DE_DRENAGEM_PLUVIAL_ROMPIDA", False),
    "INEXISTENCIA DE SISTEMA DE DRENAGEM": ("REDE_PUBLICA_DE_DRENAGEM_PLUVIAL_ROMPIDA", False),

    # variantes de digitação de "avaliação de risco"
    "AVALICAO DE RISCO": ("AVALIACAO_DE_RISCO", True),
    "AVALIACAO D ERISCO": ("AVALIACAO_DE_RISCO", True),
    "REAVALIACAO DE RISCO": ("AVALIACAO_DE_RISCO", True),

    # novos, ainda sem correspondência no catálogo
    "INFILTRACCAO": ("INFILTRACAO", True),
    "COLAPSO ESTRUTURAL": ("COLAPSO_ESTRUTURAL", True),
    "COLAPSO DE LAJE": ("COLAPSO_ESTRUTURAL", True),
    "COLAPSO DE MURO DE FECHAMENTO": ("COLAPSO_DE_MURO_DE_ARRIMO", True),
    "REMOCAO DE TERRA": ("REMOCAO_DE_TERRA", True),
    "EXPLOSAO": ("EXPLOSAO", True),
    "ACIDENTE COM VEICULO": ("ACIDENTE_COM_VEICULO", True),
    "DENUNCIA": ("DENUNCIA", True),
    "ALUGUEL SOCIAL": ("ALUGUEL_SOCIAL", True),
    "AVALIACAO PARA ALUGUEL SOCIAL": ("ALUGUEL_SOCIAL", True),
    "APROVACAO/REPROVACAO DE OBRA PROVISORIA": ("VISTORIA_DE_OBRA", True),
}

MAPA_GRAU = {
    "ALTO": "ALTO", "MEDIO": "MEDIO", "BAIXO": "BAIXO",
    "MUITO ALTO": "MUITO_ALTO", "MUTO ALTO": "MUITO_ALTO",
    # resultado legítimo de vistoria: foi ao local e não constatou risco
    "NAO CONSTATADO": "NAO_CONSTATADO", "NAO COSNTATADO": "NAO_CONSTATADO",
    "NAO CONSTADO": "NAO_CONSTATADO", "NAO CONTESTADO": "NAO_CONSTATADO",
    "NAO ATESTADO": "NAO_CONSTATADO", "NAO CONTATADO": "NAO_CONSTATADO",
}
GRAU_SEM_INFO = {"NAO ESPECIFICADO", "NAO INFORMADO", "NAO IMNFORMADO",
                 "PENDENTE", "-", "****", "******"}

MAPA_INTERDICAO = {
    "NAO": "NAO_NECESSARIA", "NAO ": "NAO_NECESSARIA", "N": "NAO_NECESSARIA",
    "SIM": "TOTAL", "PARCIAL": "PARCIAL", "SIM / PARCIAL": "PARCIAL",
    "DESOCUPACAO": "DESOCUPACAO", "SINALIZACAO": "SINALIZACAO",
}

MAPA_STATUS_VISTORIA = {"REALIZADA": "REALIZADA", "PENDENTE": "PENDENTE",
                        "DISPENSAVEL": "DISPENSAVEL"}
MAPA_STATUS_RELATORIO = {"CONCLUIDO": "CONCLUIDO", "PENDENTE": "PENDENTE",
                         "DISPENSAVEL": "DISPENSAVEL"}

# Nomes que não são vistoriadores da Defesa Civil de Sabará
NAO_VISTORIADOR = {"DEMAIS SECRETARIAS", "NOVA LIMA", "SECRETARIA DE OBRAS",
                   "DEFESA CIVIL", "SEC. OBRAS", "OBRAS"}

# Apelidos e erros de digitação com correspondência inequívoca
CORRECAO_VISTORIADOR = {
    "JOANATAS": "JONATAS", "PRICILLA": "PRISCILLA", "YASMIM": "YASMIN",
    "YASMIM RIBEIRO": "YASMIN RIBEIRO", "ROGERIO E": "ROGERIO",
    "DOUGLAS M": "DOUGLAS MARTINS", "PEDRO P": "PEDRO PAULO",
    "LEANDRO S": "LEANDRO SANTOS", "PAULO R": "PAULO ROGERIO",
    "RAFAEL A": "RAFAEL ALMEIDA", "LEANDRO JESUS": "LEANDRO DE JESUS",
}

# Primeiros nomes que aparecem sozinhos e também compostos: não dá para saber a
# qual pessoa se referem. Ficam como estão e vão para a aba de revisão.
AMBIGUOS = {"LEANDRO", "PAULO", "PEDRO", "DOUGLAS", "RAFAEL", "ROGERIO", "MARCOS"}


# ════════════════════════════════════════════════════════════════════════════
#  Leitura
# ════════════════════════════════════════════════════════════════════════════

wb = openpyxl.load_workbook(ENTRADA, data_only=True)


def ler(aba, chave_obrigatoria):
    ws = wb[aba]
    cab = [c.value for c in ws[1]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {cab[i]: row[i] for i in range(len(cab)) if cab[i]}
        if not texto(d.get(chave_obrigatoria)):
            continue
        out.append(d)
    return out


linhas = [("2025", l) for l in ler("OCORRENCIAS_2025", "N_DA_VISTORIA")] + \
         [("2026", l) for l in ler("OCORRENCIAS_2026", "N_DA_VISTORIA")]

# Aba de relatórios: enriquece com moradores, ocupação, orientações, encaminhamentos
relatorios = {}
for r in ler("RELATORIOS PRONTOS 2025", "N° OCORRENCIA"):
    relatorios[texto(r.get("N° OCORRENCIA"))] = r

registro = collections.defaultdict(list)   # trilha das decisões, para as abas de revisão


# ════════════════════════════════════════════════════════════════════════════
#  Normalizadores
# ════════════════════════════════════════════════════════════════════════════

def norm_bairro(bruto):
    b = texto(bruto)
    if not b:
        return NA, ""
    k = chave(b)
    obs = ""
    if k in ("NAO INFORMADO", "-", "N/A"):
        return NA, ""
    # Dois bairros na mesma célula: decisão do usuário é ficar com o primeiro.
    if "/" in b:
        obs = f"segundo bairro descartado: {b}"
        b = b.split("/")[0]
        k = chave(b)
    # numeral romano → arábico (Rosário I e Rosário 1 são o mesmo)
    romanos = {" I": " 1", " II": " 2", " III": " 3", " IV": " 4", " V": " 5"}
    for r, a in romanos.items():
        if k.endswith(r):
            k = k[: -len(r)] + a
            break
    canonico = titulo(k)
    registro["bairros"].append((b if not obs else texto(bruto), canonico, obs))
    return canonico, obs


def norm_vistoriadores(bruto):
    """Decisões do usuário: só os 4 primeiros entram; nomes que não são
    vistoriadores saem. Nada disso marca a linha para revisão — fica registrado
    na coluna AJUSTES_APLICADOS."""
    v = texto(bruto)
    if not v:
        return [], ""
    nomes, ajustes = [], []
    for parte in re.split(r"[,/;]|\bE\b|\+|&", v):
        p = chave(parte)
        p = re.sub(r"^(SGT|CB|SD)\.?\s+", "", p)          # patente
        p = re.sub(r"\s*\((BH|CONTAGEM|NOVA LIMA)\)\s*", "", p)
        p = re.sub(r"\s+(BH|CONTAGEM)$", "", p)
        p = p.strip(" .-")
        if not p or p in NAO_VISTORIADOR:
            # Decisão do usuário: quem não é vistoriador sai sem alarde.
            if p:
                ajustes.append(f"nome descartado (não é vistoriador): {p}")
            continue
        p = CORRECAO_VISTORIADOR.get(p, p)
        # A ambiguidade de primeiro nome ("Leandro" é o Santos ou o de Jesus?)
        # se resolve UMA vez na aba de-para, não em cada uma das 776 linhas onde
        # o nome aparece. Por isso ela não marca a linha para revisão.
        canonico = titulo(p)
        if canonico not in nomes:
            nomes.append(canonico)
        registro["vistoriadores"].append((
            texto(parte), canonico,
            "primeiro nome isolado — confirmar de qual pessoa se trata"
            if p in AMBIGUOS else ""))
    if len(nomes) > 4:
        ajustes.append(f"vistoriadores além do 4º descartados: {', '.join(nomes[4:])}")
        nomes = nomes[:4]
    return nomes, "; ".join(ajustes)


def norm_tipificacoes(bruto):
    """Uma célula pode trazer várias tipificações separadas por vírgula."""
    v = texto(bruto)
    if not v:
        return [], [], ""
    codigos, novos, obs = [], [], []
    for parte in re.split(r"[,;]", v):
        k = chave(parte).strip(" .-")
        if not k or k in ("NAO ESPECIFICADO", "NAO INFORMADO", "-"):
            continue
        achou = MAPA_TIPIFICACAO.get(k)
        if not achou:   # tenta casar por prefixo, cobrindo variações longas
            for termo, val in MAPA_TIPIFICACAO.items():
                if k.startswith(termo) or termo.startswith(k):
                    achou = val
                    break
        if achou:
            cod, eh_novo = achou
            if cod not in codigos:
                codigos.append(cod)
                if eh_novo:
                    novos.append(cod)
            registro["tipificacoes"].append((texto(parte), cod, "novo" if eh_novo else ""))
        else:
            obs.append(f"tipificação não mapeada: {texto(parte)}")
            registro["tipificacoes"].append((texto(parte), "", "NÃO MAPEADA"))
    return codigos, novos, "; ".join(obs)


def norm_grau(bruto):
    k = chave(bruto)
    if not k or k in GRAU_SEM_INFO:
        return NA
    return MAPA_GRAU.get(k, NA)


def norm_interdicao(bruto):
    k = chave(bruto)
    if not k:
        return NA
    return MAPA_INTERDICAO.get(k, NA)


def norm_sim_nao(bruto):
    k = chave(bruto)
    if k in ("SIM", "S"):
        return "SIM"
    if k in ("NAO", "N"):
        return "NAO"
    return NA


# ════════════════════════════════════════════════════════════════════════════
#  Montagem
# ════════════════════════════════════════════════════════════════════════════

COLUNAS = [
    ("PROTOCOLO", "Etapa 1"), ("N_ORIGINAL_PLANILHA", "Etapa 1"),
    ("DATA_ABERTURA", "Etapa 1"), ("HORA_ABERTURA", "Etapa 1"),
    ("SOLICITANTE_NOME", "Etapa 1"), ("SOLICITANTE_CPF", "Etapa 1"),
    ("SOLICITANTE_TELEFONE", "Etapa 1"), ("SOLICITANTE_EMAIL", "Etapa 1"),
    ("ENDERECO", "Etapa 1"), ("NUMERO", "Etapa 1"), ("BAIRRO", "Etapa 1"),
    ("CIDADE", "Etapa 1"), ("UF", "Etapa 1"),
    ("DESCRICAO_PROBLEMA", "Etapa 1"),
    ("TIPIFICACAO_INICIAL", "Etapa 2"), ("GRAU_RISCO_INICIAL", "Etapa 2"),
    ("EMERGENCIA", "Etapa 2"),
    ("DATA_AGENDAMENTO", "Etapa 3"),
    ("VISTORIADOR_1", "Etapa 3"), ("VISTORIADOR_2", "Etapa 3"),
    ("VISTORIADOR_3", "Etapa 3"), ("VISTORIADOR_4", "Etapa 3"),
    ("DATA_VISTORIA", "Etapa 4"), ("STATUS_VISTORIA", "Etapa 4"),
    ("TIPIFICACAO_VISTORIA", "Etapa 4"), ("GRAU_RISCO_ENCONTRADO", "Etapa 4"),
    ("INTERDICAO", "Etapa 4"), ("REMOCAO", "Etapa 4"),
    ("TOTAL_MORADORES", "Etapa 4"), ("REGIME_OCUPACAO", "Etapa 4"),
    ("ORIENTACOES", "Etapa 4"), ("OBSERVACOES_VISTORIA", "Etapa 4"),
    ("NOTIFICADO", "Etapa 5"), ("FORMA_RECEBIMENTO", "Etapa 5"),
    ("STATUS_RELATORIO", "Relatório"), ("DATA_RELATORIO", "Relatório"),
    ("ENCAMINHAMENTOS", "Etapa 6"), ("RETORNO_ENCAMINHAMENTOS", "Etapa 6"),
    ("STATUS_OCORRENCIA", "Sistema"),
    ("AJUSTES_APLICADOS", "Conferência"),
    ("REVISAR", "Conferência"), ("MOTIVO_REVISAO", "Conferência"),
]

saida, novos_catalogo = [], collections.Counter()
seq = collections.Counter()

descartadas = []

for ano, l in linhas:
    num = texto(l.get("N_DA_VISTORIA"))
    obs_rev, ajustes = [], []

    data_sol = como_data(l.get("DATA_SOLICITACAO"))
    if not data_sol:
        # Sem data de solicitação a linha não é uma ocorrência: na planilha são
        # números pré-impressos, com todo o resto em branco. Ficam de fora.
        preenchidos = sum(1 for k, v in l.items()
                          if k != "N_DA_VISTORIA" and texto(v))
        descartadas.append([num, ano, preenchidos,
                            "sem data de solicitação — linha em branco"
                            if preenchidos == 0 else
                            "sem data de solicitação"])
        continue

    seq[ano] += 1
    protocolo = f"{ano}-{seq[ano]:04d}"

    bairro, aj_b = norm_bairro(l.get("BAIRRO"))
    if aj_b:
        ajustes.append(aj_b)

    vist, aj_v = norm_vistoriadores(l.get("VISTORIADORES"))
    if aj_v:
        ajustes.append(aj_v)

    tips, novos, obs_t = norm_tipificacoes(l.get("TIPIFICACAO_OCORRENCIA"))
    if obs_t:
        obs_rev.append(obs_t)   # tipificação sem correspondência ainda pede olhar
    for n in novos:
        novos_catalogo[n] += 1

    data_vist = como_data(l.get("DATA DA VISTORIA"))
    if data_vist and not (0 <= (data_vist - data_sol).days <= 365):
        ajustes.append(f"data de vistoria descartada por inconsistência ({data_vist})")
        data_vist = None

    rel = relatorios.get(num, {})
    moradores = so_digitos(rel.get("N° DE MORADORES"))

    cpf = so_digitos(l.get("CPF/IDENTIDADE"))
    if cpf and len(cpf) != 11:
        ajustes.append(f"documento descartado — {len(cpf)} dígitos, não é CPF")
        cpf = ""

    status_v = MAPA_STATUS_VISTORIA.get(chave(l.get("STATUS_VISTORIA")), NA)
    status_r = MAPA_STATUS_RELATORIO.get(chave(l.get("STATUS_RELATORIO")), NA)
    if status_v == "REALIZADA" and status_r in ("CONCLUIDO", "DISPENSAVEL"):
        status_oc = "ENCERRADA"
    elif status_v == "REALIZADA":
        status_oc = "VISTORIA_REALIZADA"
    elif status_v == "DISPENSAVEL":
        status_oc = "ENCERRADA"
    else:
        status_oc = "VISTORIA_SOLICITADA"

    obs_texto = " | ".join(filter(None, [
        texto(l.get("OBSERVAÇÃO")),
        f"Despachos: {texto(l.get('DESPACHOS ENVIADOS'))}" if texto(l.get("DESPACHOS ENVIADOS")) else "",
        f"Resposta: {texto(l.get('RESPOSTA DESPACHOS'))}" if texto(l.get("RESPOSTA DESPACHOS")) else "",
        f"Conclusão do relatório: {texto(rel.get('CONCLUSÃO'))}" if texto(rel.get("CONCLUSÃO")) else "",
    ]))

    linha = {
        "PROTOCOLO": protocolo,
        "N_ORIGINAL_PLANILHA": num or NA,
        "DATA_ABERTURA": data_sol.strftime("%d/%m/%Y") if data_sol else NA,
        "HORA_ABERTURA": como_hora(l.get("HORARIO")) or NA,
        "SOLICITANTE_NOME": titulo(l.get("NOME DO SOLICITANTE")) or NA,
        "SOLICITANTE_CPF": cpf or NA,
        "SOLICITANTE_TELEFONE": texto(l.get("TELEFONE")) or NA,
        "SOLICITANTE_EMAIL": texto(l.get("EMAIL")).lower() or NA,
        "ENDERECO": titulo(l.get("ENDEREÇO")) or NA,
        "NUMERO": texto(l.get("N°")) or NA,
        "BAIRRO": bairro,
        "CIDADE": "Sabará",
        "UF": "MG",
        "DESCRICAO_PROBLEMA": texto(l.get("DESCRIÇÃO PRELIMINAR")) or texto(l.get("TIPIFICACAO_OCORRENCIA")) or NA,
        "TIPIFICACAO_INICIAL": "; ".join(tips) if tips else NA,
        "GRAU_RISCO_INICIAL": NA,
        "EMERGENCIA": "SIM" if any("EMERGENCIA" in t for t in tips) else "NAO",
        "DATA_AGENDAMENTO": (como_data(l.get("ABERTURA DA VISTORIA")) or data_vist or "") and
                            (como_data(l.get("ABERTURA DA VISTORIA")) or data_vist).strftime("%d/%m/%Y") or NA,
        "VISTORIADOR_1": vist[0] if len(vist) > 0 else NA,
        "VISTORIADOR_2": vist[1] if len(vist) > 1 else NA,
        "VISTORIADOR_3": vist[2] if len(vist) > 2 else NA,
        "VISTORIADOR_4": vist[3] if len(vist) > 3 else NA,
        "DATA_VISTORIA": data_vist.strftime("%d/%m/%Y") if data_vist else NA,
        "STATUS_VISTORIA": status_v,
        "TIPIFICACAO_VISTORIA": "; ".join(tips) if tips else NA,
        "GRAU_RISCO_ENCONTRADO": norm_grau(l.get("GRAU_RISCO")),
        "INTERDICAO": norm_interdicao(l.get("INTERDIÇÃO")),
        "REMOCAO": NA,
        "TOTAL_MORADORES": moradores or NA,
        "REGIME_OCUPACAO": titulo(rel.get("OCUPAÇÃO DO IMOVEL")) or NA,
        "ORIENTACOES": texto(rel.get("ORIENTAÇÕES")) or NA,
        "OBSERVACOES_VISTORIA": obs_texto or NA,
        "NOTIFICADO": norm_sim_nao(l.get("NOTIFICAÇÃO")),
        "FORMA_RECEBIMENTO": (lambda f: "EMAIL" if "MAIL" in chave(f) else
                              ("PRESENCIAL" if f else NA))(texto(rel.get("FORMA DE ENTREGA DO RELATORIO"))),
        "STATUS_RELATORIO": status_r,
        "DATA_RELATORIO": (lambda d: d.strftime("%d/%m/%Y") if d else NA)(como_data(l.get("DATA_RELATORIO"))),
        "ENCAMINHAMENTOS": texto(rel.get("ENCAMINHAMENTOS")) or NA,
        "RETORNO_ENCAMINHAMENTOS": texto(l.get("RESPOSTA DESPACHOS")) or NA,
        "STATUS_OCORRENCIA": status_oc,
        # Registro do que foi resolvido automaticamente, conforme as decisões
        # tomadas. Não pede ação — existe para permitir auditoria.
        "AJUSTES_APLICADOS": " | ".join(ajustes) if ajustes else NA,
        "REVISAR": "SIM" if obs_rev else "NAO",
        "MOTIVO_REVISAO": " | ".join(obs_rev) if obs_rev else NA,
    }
    saida.append(linha)


# ════════════════════════════════════════════════════════════════════════════
#  Escrita
# ════════════════════════════════════════════════════════════════════════════

out = openpyxl.Workbook()
CAB = Font(bold=True, color="FFFFFF")
FUNDO = PatternFill("solid", fgColor="2A78D6")
ALERTA = PatternFill("solid", fgColor="FDE9E9")


def escrever_aba(ws, cabecalho, linhas_, larguras=None):
    ws.append(cabecalho)
    for c in ws[1]:
        c.font = CAB
        c.fill = FUNDO
        c.alignment = Alignment(vertical="center")
    for l in linhas_:
        ws.append(l)
    ws.freeze_panes = "A2"
    for i, h in enumerate(cabecalho, 1):
        larg = (larguras or {}).get(h, min(max(len(str(h)) + 3, 14), 42))
        ws.column_dimensions[get_column_letter(i)].width = larg
    ws.auto_filter.ref = ws.dimensions


# — aba principal
ws = out.active
ws.title = "OCORRENCIAS"
escrever_aba(ws, [c for c, _ in COLUNAS],
             [[l[c] for c, _ in COLUNAS] for l in saida])
col_rev = [c for c, _ in COLUNAS].index("REVISAR") + 1
for r in range(2, ws.max_row + 1):
    if ws.cell(r, col_rev).value == "SIM":
        ws.cell(r, col_rev).fill = ALERTA

# — de-para
def dedup(chave_reg):
    c = collections.Counter()
    marca = {}
    for orig, canon, obs in registro[chave_reg]:
        c[(orig, canon)] += 1
        if obs:
            marca[(orig, canon)] = obs
    return [[o, cn, n, marca.get((o, cn), "")] for (o, cn), n in c.most_common()]

escrever_aba(out.create_sheet("DE-PARA VISTORIADORES"),
             ["NOME NA PLANILHA", "NOME NORMALIZADO", "OCORRÊNCIAS", "OBSERVAÇÃO"],
             dedup("vistoriadores"), {"NOME NA PLANILHA": 28, "NOME NORMALIZADO": 28, "OBSERVAÇÃO": 30})

escrever_aba(out.create_sheet("DE-PARA BAIRROS"),
             ["BAIRRO NA PLANILHA", "BAIRRO NORMALIZADO", "OCORRÊNCIAS", "OBSERVAÇÃO"],
             dedup("bairros"), {"BAIRRO NA PLANILHA": 34, "BAIRRO NORMALIZADO": 34, "OBSERVAÇÃO": 40})

escrever_aba(out.create_sheet("DE-PARA TIPIFICACOES"),
             ["TIPIFICAÇÃO NA PLANILHA", "CÓDIGO NO SISTEMA", "OCORRÊNCIAS", "SITUAÇÃO"],
             dedup("tipificacoes"), {"TIPIFICAÇÃO NA PLANILHA": 46, "CÓDIGO NO SISTEMA": 40})

# — catálogo a criar
escrever_aba(out.create_sheet("CATALOGO A CRIAR"),
             ["CÓDIGO", "OCORRÊNCIAS", "OBSERVAÇÃO"],
             [[c, n, "criar em Opções de campo da vistoria antes de importar"]
              for c, n in novos_catalogo.most_common()],
             {"CÓDIGO": 40, "OBSERVAÇÃO": 52})

# — pendências (o que ainda pede decisão humana)
pend = [[l["PROTOCOLO"], l["N_ORIGINAL_PLANILHA"], l["MOTIVO_REVISAO"]]
        for l in saida if l["REVISAR"] == "SIM"]
escrever_aba(out.create_sheet("PENDENCIAS"),
             ["PROTOCOLO", "N° NA PLANILHA", "MOTIVO"], pend,
             {"MOTIVO": 90})

# — ajustes automáticos (rastreabilidade das decisões já tomadas)
aj = [[l["PROTOCOLO"], l["N_ORIGINAL_PLANILHA"], l["AJUSTES_APLICADOS"]]
      for l in saida if l["AJUSTES_APLICADOS"] != NA]
escrever_aba(out.create_sheet("AJUSTES APLICADOS"),
             ["PROTOCOLO", "N° NA PLANILHA", "O QUE FOI AJUSTADO"], aj,
             {"O QUE FOI AJUSTADO": 90})

# — linhas descartadas
escrever_aba(out.create_sheet("LINHAS DESCARTADAS"),
             ["N° NA PLANILHA", "ABA", "CAMPOS PREENCHIDOS", "MOTIVO"],
             descartadas, {"MOTIVO": 46})

out.save(SAIDA)

# ── Resumo ──────────────────────────────────────────────────────────────────
print(f"Gerado: {SAIDA}")
print(f"  ocorrências normalizadas : {len(saida)}")
print(f"  linhas descartadas       : {len(descartadas)}")
print(f"  com ajuste automático    : {sum(1 for l in saida if l['AJUSTES_APLICADOS']!=NA)}")
print(f"  marcadas para revisão    : {sum(1 for l in saida if l['REVISAR']=='SIM')}")
print(f"  vistoriadores distintos  : {len({c for _,c,_ in registro['vistoriadores']})}")
print(f"  bairros distintos        : {len({c for _,c,_ in registro['bairros']})}")
print(f"  tipificações distintas   : {len({c for _,c,_ in registro['tipificacoes'] if c})}")
print(f"  opções novas de catálogo : {len(novos_catalogo)}")
nm = [o for o, c, s in registro["tipificacoes"] if s == "NÃO MAPEADA"]
print(f"  tipificações não mapeadas: {len(set(nm))} -> {sorted(set(nm))[:6]}")
